import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

src = r'C:\Users\santi\Downloads\Santiago Personal Goal Tracker (10).xlsx'
out = r'C:\Users\santi\Downloads\Santiago Goal Tracker Dashboard.xlsx'

wb = load_workbook(src)

# ── Color palette ──
DARK_BG = "1B1F2B"
CARD_BG = "252A3A"
ACCENT_BLUE = "4A90D9"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_ORANGE = "F39C12"
ACCENT_PURPLE = "9B59B6"
ACCENT_TEAL = "1ABC9C"
WHITE = "FFFFFF"
LIGHT_GRAY = "BDC3C7"
GOLD_COLOR = "F1C40F"

title_font = Font(name="Arial", size=20, bold=True, color=WHITE)
section_font = Font(name="Arial", size=14, bold=True, color=ACCENT_BLUE)
header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
data_font = Font(name="Arial", size=10, color=LIGHT_GRAY)
value_font = Font(name="Arial", size=22, bold=True, color=WHITE)
label_font = Font(name="Arial", size=9, color=LIGHT_GRAY)
kpi_label_font = Font(name="Arial", size=9, color=LIGHT_GRAY)
kpi_value_font = Font(name="Arial", size=18, bold=True, color=WHITE)
small_font = Font(name="Arial", size=8, color=LIGHT_GRAY)

dark_fill = PatternFill("solid", fgColor=DARK_BG)
card_fill = PatternFill("solid", fgColor=CARD_BG)
blue_fill = PatternFill("solid", fgColor=ACCENT_BLUE)
green_fill = PatternFill("solid", fgColor=ACCENT_GREEN)
red_fill = PatternFill("solid", fgColor=ACCENT_RED)
orange_fill = PatternFill("solid", fgColor=ACCENT_ORANGE)
purple_fill = PatternFill("solid", fgColor=ACCENT_PURPLE)
teal_fill = PatternFill("solid", fgColor=ACCENT_TEAL)
gold_fill = PatternFill("solid", fgColor=GOLD_COLOR)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="3A3F55"),
    right=Side(style="thin", color="3A3F55"),
    top=Side(style="thin", color="3A3F55"),
    bottom=Side(style="thin", color="3A3F55")
)

def style_range(ws, r1, c1, r2, c2, fill=None, font=None, alignment=None, border=None):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(row=r, column=c)
            if fill: cell.fill = fill
            if font: cell.font = font
            if alignment: cell.alignment = alignment
            if border: cell.border = border

# ═══════════════════════════════════════════
# DASHBOARD SHEET
# ═══════════════════════════════════════════
ds = wb.create_sheet("Dashboard", 0)
ds.sheet_properties.tabColor = ACCENT_BLUE

# Set column widths
for c in range(1, 30):
    ds.column_dimensions[get_column_letter(c)].width = 14
ds.column_dimensions['A'].width = 3

# Dark background for entire visible area
style_range(ds, 1, 1, 100, 29, fill=dark_fill)

# ── Title ──
ds.merge_cells('B2:N2')
ds['B2'] = "SANTIAGO'S PERSONAL GOAL TRACKER"
ds['B2'].font = Font(name="Arial", size=24, bold=True, color=WHITE)
ds['B2'].alignment = Alignment(horizontal="left", vertical="center")
ds['B2'].fill = dark_fill

ds.merge_cells('B3:N3')
ds['B3'] = "Life Dashboard  |  2024-2025 Progress Overview"
ds['B3'].font = Font(name="Arial", size=11, color=LIGHT_GRAY)
ds['B3'].alignment = Alignment(horizontal="left", vertical="center")

# ═══════════════════════════════════════════
# SECTION 1: HEALTH KPI CARDS (Row 5-9)
# ═══════════════════════════════════════════
ds.merge_cells('B5:N5')
ds['B5'] = "HEALTH METRICS"
ds['B5'].font = Font(name="Arial", size=14, bold=True, color=ACCENT_GREEN)
ds['B5'].fill = dark_fill

# Health data 2025 (latest available: Nov 2025 = col index 10 in 2025 block)
# 2025 months: Jan=row28 block, columns shift by 3 per month
# Latest with data: Nov 2025
health_2025 = {
    'months': ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov'],
    'weight': [139.4, 139.6, 138, 140.8, 138.8, 135.2, 135.4, 139.8, 144.2, 149.2, 151.2],
    'body_fat': [21.5, 21.5, 21, 21.8, 21.3, 20.2, 20.3, 21.5, 22.7, 24.1, 24.6],
    'muscle': [104, 104, 103.6, 104.4, 103.8, 102.6, 102.6, 104.2, 105.8, 107.6, 108.2],
    'avg_steps': [6772, 6713, 7384, 7260, 7189, 8003, 8373, 7857, 12449, 7648, 5851],
    'total_steps': [217000, 187981, 228910, 217815, 222878, 240101, 259569, 243573, 373485, 237118, 76074],
    'kms': [132.8, 121.5, 146.8, 140.4, 142.9, 153.9, 170.8, 157.3, 246.5, 156.7, 48.7],
    'body_water': [56.7, 56.6, 57, 56.4, 56.8, 57.6, 57.6, 56.6, 55.8, 54.8, 54.5],
    'metabolic_age': [26, 26, 25, 26, 26, 24, 25, 26, 27, 30, 30],
}

# KPI Cards Row 7-9
kpi_data = [
    ("Current Weight", f"{health_2025['weight'][-1]} lbs", green_fill),
    ("Body Fat %", f"{health_2025['body_fat'][-1]}%", orange_fill),
    ("Muscle Mass", f"{health_2025['muscle'][-1]} lbs", blue_fill),
    ("Metabolic Age", f"{health_2025['metabolic_age'][-1]}", purple_fill),
    ("Avg Steps (Nov)", f"{health_2025['avg_steps'][-1]:,}", teal_fill),
    ("Total Kms YTD", f"{sum(health_2025['kms']):.1f}", red_fill),
]

col = 2
for label, value, fill_color in kpi_data:
    ds.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    ds.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
    style_range(ds, 6, col, 9, col+1, fill=card_fill, border=thin_border)
    # Color strip at top
    style_range(ds, 6, col, 6, col+1, fill=fill_color)
    ds.cell(row=7, column=col, value=label).font = kpi_label_font
    ds.cell(row=7, column=col).alignment = center
    ds.cell(row=8, column=col, value=value).font = kpi_value_font
    ds.cell(row=8, column=col).alignment = center
    col += 2

# ── HEALTH DATA TABLE for charts (Row 11+) ──
ds['B11'] = "Month"
ds['B11'].font = header_font
ds['B11'].fill = blue_fill
ds['B11'].alignment = center

headers = ["Weight (lbs)", "Body Fat %", "Muscle Mass", "Avg Steps", "Kms", "Body Water %", "Metabolic Age"]
for i, h in enumerate(headers):
    cell = ds.cell(row=11, column=3+i, value=h)
    cell.font = header_font
    cell.fill = blue_fill
    cell.alignment = center

for idx, month in enumerate(health_2025['months']):
    r = 12 + idx
    ds.cell(row=r, column=2, value=month).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center

    vals = [
        health_2025['weight'][idx], health_2025['body_fat'][idx],
        health_2025['muscle'][idx], health_2025['avg_steps'][idx],
        health_2025['kms'][idx], health_2025['body_water'][idx],
        health_2025['metabolic_age'][idx]
    ]
    for j, v in enumerate(vals):
        cell = ds.cell(row=r, column=3+j, value=v)
        cell.font = data_font
        cell.fill = card_fill
        cell.alignment = center
        cell.number_format = '#,##0.0' if j < 3 or j == 4 or j == 5 else '#,##0'

last_health_row = 12 + len(health_2025['months']) - 1

# ── Chart 1: Weight & Body Fat Trend ──
c1 = LineChart()
c1.title = "Weight & Body Fat Trend (2025)"
c1.style = 10
c1.height = 14
c1.width = 22
c1.y_axis.title = "Weight (lbs)"
cats = Reference(ds, min_col=2, min_row=12, max_row=last_health_row)
weight_data = Reference(ds, min_col=3, min_row=11, max_row=last_health_row)
bf_data = Reference(ds, min_col=4, min_row=11, max_row=last_health_row)
c1.add_data(weight_data, titles_from_data=True)
c1.add_data(bf_data, titles_from_data=True)
c1.set_categories(cats)
s1 = c1.series[0]
s1.graphicalProperties.line.width = 28000
s2 = c1.series[1]
s2.graphicalProperties.line.width = 28000
ds.add_chart(c1, "B24")

# ── Chart 2: Steps & Kms ──
c2 = BarChart()
c2.type = "col"
c2.title = "Monthly Steps (2025)"
c2.style = 10
c2.height = 14
c2.width = 22
steps_data = Reference(ds, min_col=6, min_row=11, max_row=last_health_row)
c2.add_data(steps_data, titles_from_data=True)
c2.set_categories(cats)
c2.series[0].graphicalProperties.solidFill = ACCENT_TEAL
ds.add_chart(c2, "B40")

# ── Chart 3: Muscle Mass Trend ──
c3 = LineChart()
c3.title = "Muscle Mass Trend (2025)"
c3.style = 10
c3.height = 14
c3.width = 22
muscle_data = Reference(ds, min_col=5, min_row=11, max_row=last_health_row)
c3.add_data(muscle_data, titles_from_data=True)
c3.set_categories(cats)
c3.series[0].graphicalProperties.line.width = 28000
ds.add_chart(c3, "L24")

# ═══════════════════════════════════════════
# SECTION 2: FINANCIALS (Row 56+)
# ═══════════════════════════════════════════
fin_row = 56
ds.merge_cells(f'B{fin_row}:N{fin_row}')
ds[f'B{fin_row}'] = "FINANCIAL OVERVIEW"
ds[f'B{fin_row}'].font = Font(name="Arial", size=14, bold=True, color=GOLD_COLOR)
ds[f'B{fin_row}'].fill = dark_fill

# 2025 financial data
fin_2025 = {
    'months': ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'],
    'total_cad': [323406.51, 226334.28, 230154.51, 226300, 220550, 231129.49, 313219.55, 286018.74, 293952.39, 260776.35, 242610.16, 205171.62],
    'total_usd': [218511.28, 156149.28, 160326.99, 164110, 160777, 169900.73, 226450.86, 208348.09, 211227.93, 186240.22, 173935.97, 150265.78],
    'crypto_cad': [315836.5, 219259.5, 220123.42, 214000, 210795, 219000, 302174, 273425, 278767, 240310.56, 222606, 177401],
    'debt_cad': [17583.04, 14974.19, 13330.24, 11828, 11031, 9835.78, 12988.88, 10015.54, 11934.19, 9675.05, 10757, 9428],
    'cash_cad': [6225.05, 3069.2, 2856.23, 4310, 1396, 20, 1400, 15, 70, 1460, 0, 0],
}

# Financial KPI cards
fin_kpi_row = fin_row + 2
fkpis = [
    ("Net Worth (CAD)", f"${fin_2025['total_cad'][-1]:,.0f}", red_fill),
    ("Net Worth (USD)", f"${fin_2025['total_usd'][-1]:,.0f}", orange_fill),
    ("Crypto (CAD)", f"${fin_2025['crypto_cad'][-1]:,.0f}", purple_fill),
    ("Total Debt (CAD)", f"${fin_2025['debt_cad'][-1]:,.0f}", blue_fill),
    ("YTD Change CAD", f"-$118,235", red_fill),
    ("YTD Change USD", f"-$68,246", red_fill),
]
col = 2
for label, value, fill_color in fkpis:
    ds.merge_cells(start_row=fin_kpi_row, start_column=col, end_row=fin_kpi_row, end_column=col+1)
    ds.merge_cells(start_row=fin_kpi_row+1, start_column=col, end_row=fin_kpi_row+1, end_column=col+1)
    style_range(ds, fin_kpi_row-1, col, fin_kpi_row+2, col+1, fill=card_fill, border=thin_border)
    style_range(ds, fin_kpi_row-1, col, fin_kpi_row-1, col+1, fill=fill_color)
    ds.cell(row=fin_kpi_row, column=col, value=label).font = kpi_label_font
    ds.cell(row=fin_kpi_row, column=col).alignment = center
    ds.cell(row=fin_kpi_row+1, column=col, value=value).font = kpi_value_font
    ds.cell(row=fin_kpi_row+1, column=col).alignment = center
    col += 2

# Financial data table
ftbl_row = fin_kpi_row + 4
ds.cell(row=ftbl_row, column=2, value="Month").font = header_font
ds.cell(row=ftbl_row, column=2).fill = PatternFill("solid", fgColor=GOLD_COLOR)
ds.cell(row=ftbl_row, column=2).alignment = center
fin_headers = ["Net Worth CAD", "Net Worth USD", "Crypto CAD", "Debt CAD", "Cash CAD"]
for i, h in enumerate(fin_headers):
    cell = ds.cell(row=ftbl_row, column=3+i, value=h)
    cell.font = Font(name="Arial", size=10, bold=True, color=DARK_BG)
    cell.fill = PatternFill("solid", fgColor=GOLD_COLOR)
    cell.alignment = center

for idx, month in enumerate(fin_2025['months']):
    r = ftbl_row + 1 + idx
    ds.cell(row=r, column=2, value=month).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    vals = [fin_2025['total_cad'][idx], fin_2025['total_usd'][idx], fin_2025['crypto_cad'][idx], fin_2025['debt_cad'][idx], fin_2025['cash_cad'][idx]]
    for j, v in enumerate(vals):
        cell = ds.cell(row=r, column=3+j, value=v)
        cell.font = data_font
        cell.fill = card_fill
        cell.alignment = center
        cell.number_format = '$#,##0'

fin_last_row = ftbl_row + len(fin_2025['months'])

# Chart: Net Worth Trend
c4 = LineChart()
c4.title = "Net Worth Trend 2025 (CAD & USD)"
c4.style = 10
c4.height = 14
c4.width = 22
fin_cats = Reference(ds, min_col=2, min_row=ftbl_row+1, max_row=fin_last_row)
nw_cad = Reference(ds, min_col=3, min_row=ftbl_row, max_row=fin_last_row)
nw_usd = Reference(ds, min_col=4, min_row=ftbl_row, max_row=fin_last_row)
c4.add_data(nw_cad, titles_from_data=True)
c4.add_data(nw_usd, titles_from_data=True)
c4.set_categories(fin_cats)
c4.series[0].graphicalProperties.line.width = 28000
c4.series[1].graphicalProperties.line.width = 28000
ds.add_chart(c4, f"B{fin_last_row + 2}")

# Chart: Asset Breakdown (latest month pie)
c5 = PieChart()
c5.title = "Asset Allocation (Dec 2025 - CAD)"
c5.style = 10
c5.height = 14
c5.width = 16

# Put pie data in cells
pie_row = fin_last_row + 2
ds.cell(row=pie_row, column=11, value="Asset").font = header_font
ds.cell(row=pie_row, column=12, value="Value").font = header_font
assets = [("Crypto", 177401), ("Silver", 26034.62), ("Gold", 11164), ("Cash", 0)]
for i, (name, val) in enumerate(assets):
    ds.cell(row=pie_row+1+i, column=11, value=name).font = data_font
    ds.cell(row=pie_row+1+i, column=12, value=val).font = data_font

pie_cats = Reference(ds, min_col=11, min_row=pie_row+1, max_row=pie_row+len(assets))
pie_vals = Reference(ds, min_col=12, min_row=pie_row, max_row=pie_row+len(assets))
c5.add_data(pie_vals, titles_from_data=True)
c5.set_categories(pie_cats)
c5.dataLabels = DataLabelList()
c5.dataLabels.showPercent = True
c5.dataLabels.showCatName = True
ds.add_chart(c5, f"L{fin_last_row + 2}")

# ═══════════════════════════════════════════
# SECTION 3: JOB PERFORMANCE (Row 90+)
# ═══════════════════════════════════════════
job_row = 90
ds.merge_cells(f'B{job_row}:N{job_row}')
ds[f'B{job_row}'] = "JOB PERFORMANCE"
ds[f'B{job_row}'].font = Font(name="Arial", size=14, bold=True, color=ACCENT_ORANGE)
ds[f'B{job_row}'].fill = dark_fill

# 2025 AHT data
job_2025 = {
    'months': ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov'],
    'aht_actual': [670, 642, 657, 669, 694, 692, 696, 688, 681, 778, 713],
    'aht_target': [760, 750, 750, 750, 710, 710, 710, 710, 710, 710, 745],
}

jtbl_row = job_row + 2
ds.cell(row=jtbl_row, column=2, value="Month").font = header_font
ds.cell(row=jtbl_row, column=2).fill = orange_fill
ds.cell(row=jtbl_row, column=2).alignment = center
for i, h in enumerate(["AHT Actual", "AHT Target"]):
    cell = ds.cell(row=jtbl_row, column=3+i, value=h)
    cell.font = Font(name="Arial", size=10, bold=True, color=DARK_BG)
    cell.fill = orange_fill
    cell.alignment = center

for idx, month in enumerate(job_2025['months']):
    r = jtbl_row + 1 + idx
    ds.cell(row=r, column=2, value=month).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    ds.cell(row=r, column=3, value=job_2025['aht_actual'][idx]).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = center
    ds.cell(row=r, column=4, value=job_2025['aht_target'][idx]).font = data_font
    ds.cell(row=r, column=4).fill = card_fill
    ds.cell(row=r, column=4).alignment = center

job_last_row = jtbl_row + len(job_2025['months'])

c6 = BarChart()
c6.type = "col"
c6.title = "AHT: Actual vs Target (2025)"
c6.style = 10
c6.height = 14
c6.width = 22
job_cats = Reference(ds, min_col=2, min_row=jtbl_row+1, max_row=job_last_row)
aht_act = Reference(ds, min_col=3, min_row=jtbl_row, max_row=job_last_row)
aht_tgt = Reference(ds, min_col=4, min_row=jtbl_row, max_row=job_last_row)
c6.add_data(aht_act, titles_from_data=True)
c6.add_data(aht_tgt, titles_from_data=True)
c6.set_categories(job_cats)
c6.series[0].graphicalProperties.solidFill = ACCENT_GREEN
c6.series[1].graphicalProperties.solidFill = ACCENT_RED
ds.add_chart(c6, f"B{job_last_row + 2}")

# ═══════════════════════════════════════════
# SECTION 4: PERSONAL DEVELOPMENT SUMMARY
# ═══════════════════════════════════════════
pd_row = job_last_row + 20
ds.merge_cells(f'B{pd_row}:N{pd_row}')
ds[f'B{pd_row}'] = "PERSONAL DEVELOPMENT - DAY QUALITY TRACKER"
ds[f'B{pd_row}'].font = Font(name="Arial", size=14, bold=True, color=ACCENT_PURPLE)
ds[f'B{pd_row}'].fill = dark_fill

# 2025 day quality data
day_data_2025 = {
    'Green Days': 269,
    'Yellow Days': 31,
    'Pink Days': 28,
    'Blue Days': 37,
    'Sick Days': 11,
}
total_days_2025 = sum(day_data_2025.values()) - day_data_2025.get('Sick Days', 0)

pd_tbl_row = pd_row + 2
ds.cell(row=pd_tbl_row, column=2, value="Category").font = header_font
ds.cell(row=pd_tbl_row, column=2).fill = purple_fill
ds.cell(row=pd_tbl_row, column=2).alignment = center
ds.cell(row=pd_tbl_row, column=3, value="Days").font = header_font
ds.cell(row=pd_tbl_row, column=3).fill = purple_fill
ds.cell(row=pd_tbl_row, column=3).alignment = center
ds.cell(row=pd_tbl_row, column=4, value="% of Year").font = header_font
ds.cell(row=pd_tbl_row, column=4).fill = purple_fill
ds.cell(row=pd_tbl_row, column=4).alignment = center

colors_map = {
    'Green Days': ACCENT_GREEN,
    'Yellow Days': GOLD_COLOR,
    'Pink Days': "E91E63",
    'Blue Days': ACCENT_BLUE,
    'Sick Days': LIGHT_GRAY,
}

for i, (cat, days) in enumerate(day_data_2025.items()):
    r = pd_tbl_row + 1 + i
    ds.cell(row=r, column=2, value=cat).font = Font(name="Arial", size=10, bold=True, color=colors_map[cat])
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    ds.cell(row=r, column=3, value=days).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = center
    pct = days / 365
    ds.cell(row=r, column=4, value=pct).font = data_font
    ds.cell(row=r, column=4).fill = card_fill
    ds.cell(row=r, column=4).alignment = center
    ds.cell(row=r, column=4).number_format = '0.0%'

# Pie chart for day quality
c7 = PieChart()
c7.title = "Day Quality Distribution (2025)"
c7.style = 10
c7.height = 14
c7.width = 16
pd_cats = Reference(ds, min_col=2, min_row=pd_tbl_row+1, max_row=pd_tbl_row+len(day_data_2025))
pd_vals = Reference(ds, min_col=3, min_row=pd_tbl_row, max_row=pd_tbl_row+len(day_data_2025))
c7.add_data(pd_vals, titles_from_data=True)
c7.set_categories(pd_cats)
c7.dataLabels = DataLabelList()
c7.dataLabels.showPercent = True
c7.dataLabels.showCatName = True

# Color the pie slices
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
pie_colors = [ACCENT_GREEN, GOLD_COLOR, "E91E63", ACCENT_BLUE, LIGHT_GRAY]
for idx_c, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx_c)
    pt.graphicalProperties.solidFill = color
    c7.series[0].data_points.append(pt)

ds.add_chart(c7, f"F{pd_tbl_row}")

# 2024 comparison
day_data_2024 = {
    'Green Days': 266,
    'Yellow Days': 22,
    'Pink Days': 39,
    'Blue Days': 38,
    'Sick Days': 15,
}

comp_row = pd_tbl_row + len(day_data_2025) + 2
ds.cell(row=comp_row, column=2, value="Year-over-Year Comparison").font = Font(name="Arial", size=11, bold=True, color=WHITE)
ds.cell(row=comp_row, column=2).fill = card_fill
ds.merge_cells(start_row=comp_row, start_column=2, end_row=comp_row, end_column=5)

comp_hdr = comp_row + 1
for i, h in enumerate(["Category", "2024", "2025", "Change"]):
    cell = ds.cell(row=comp_hdr, column=2+i, value=h)
    cell.font = header_font
    cell.fill = purple_fill
    cell.alignment = center

for i, cat in enumerate(day_data_2025.keys()):
    r = comp_hdr + 1 + i
    ds.cell(row=r, column=2, value=cat).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=3, value=day_data_2024[cat]).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = center
    ds.cell(row=r, column=4, value=day_data_2025[cat]).font = data_font
    ds.cell(row=r, column=4).fill = card_fill
    ds.cell(row=r, column=4).alignment = center
    change = day_data_2025[cat] - day_data_2024[cat]
    cell = ds.cell(row=r, column=5, value=change)
    cell.font = Font(name="Arial", size=10, bold=True, color=ACCENT_GREEN if (change > 0 and cat == 'Green Days') or (change < 0 and cat != 'Green Days') else ACCENT_RED)
    cell.fill = card_fill
    cell.alignment = center

# ═══════════════════════════════════════════
# SECTION 5: RELATIONSHIP MILESTONES TIMELINE
# ═══════════════════════════════════════════
rel_row = comp_hdr + len(day_data_2025) + 4
ds.merge_cells(f'B{rel_row}:N{rel_row}')
ds[f'B{rel_row}'] = "RELATIONSHIP MILESTONES (2025)"
ds[f'B{rel_row}'].font = Font(name="Arial", size=14, bold=True, color="E91E63")
ds[f'B{rel_row}'].fill = dark_fill

rel_months = ['Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
rel_highlights = [
    "Became Boyfriends, Valentine's",
    "Sunrise Port Stanley, Salsa Dance",
    "Jazz Concert, Candle Making",
    "Spain University, First Wedding",
    "100km Race, First BBQ, Fire Pit",
    "5km Run, Pool Party, Boat Shopping",
    "1st Camping, Porsche Experience",
    "Spain Travel, Alhambra, Flamenco",
    "Birthday Surprises, Online Date",
    "1 Year Anniversary, Gifted Vynil",
    "First Christmas, Avatar 3, Church"
]

rel_tbl_row = rel_row + 1
ds.cell(row=rel_tbl_row, column=2, value="Month").font = header_font
ds.cell(row=rel_tbl_row, column=2).fill = PatternFill("solid", fgColor="E91E63")
ds.cell(row=rel_tbl_row, column=2).alignment = center
ds.cell(row=rel_tbl_row, column=3, value="Key Moments").font = header_font
ds.cell(row=rel_tbl_row, column=3).fill = PatternFill("solid", fgColor="E91E63")
ds.cell(row=rel_tbl_row, column=3).alignment = center
ds.merge_cells(start_row=rel_tbl_row, start_column=3, end_row=rel_tbl_row, end_column=8)

for i, (m, h) in enumerate(zip(rel_months, rel_highlights)):
    r = rel_tbl_row + 1 + i
    ds.cell(row=r, column=2, value=m).font = Font(name="Arial", size=10, bold=True, color="E91E63")
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    ds.cell(row=r, column=2).border = thin_border
    ds.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ds.cell(row=r, column=3, value=h).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = left_align
    ds.cell(row=r, column=3).border = thin_border

# ═══════════════════════════════════════════
# SECTION 6: HEALTH 2024 DATA (for comparison chart)
# ═══════════════════════════════════════════
health_2024 = {
    'months': ['May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'],
    'weight': [152.8, 150.2, 146.8, 145.2, 148.8, 145.8, 145, 142.4],
    'body_fat': [24.8, 24.2, 23.3, 22.9, 23.8, 23.1, 22.9, 22.3],
}

# Comparison chart: Weight 2024 vs 2025
compare_row = rel_tbl_row + len(rel_months) + 4
ds.merge_cells(f'B{compare_row}:N{compare_row}')
ds[f'B{compare_row}'] = "WEIGHT JOURNEY: 2024 vs 2025"
ds[f'B{compare_row}'].font = Font(name="Arial", size=14, bold=True, color=ACCENT_TEAL)
ds[f'B{compare_row}'].fill = dark_fill

# Build a combined timeline
all_months_order = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
w24_map = dict(zip(health_2024['months'], health_2024['weight']))
w25_map = dict(zip(health_2025['months'], health_2025['weight']))

ctbl_row = compare_row + 1
ds.cell(row=ctbl_row, column=2, value="Month").font = header_font
ds.cell(row=ctbl_row, column=2).fill = teal_fill
ds.cell(row=ctbl_row, column=2).alignment = center
ds.cell(row=ctbl_row, column=3, value="2024 Weight").font = header_font
ds.cell(row=ctbl_row, column=3).fill = teal_fill
ds.cell(row=ctbl_row, column=3).alignment = center
ds.cell(row=ctbl_row, column=4, value="2025 Weight").font = header_font
ds.cell(row=ctbl_row, column=4).fill = teal_fill
ds.cell(row=ctbl_row, column=4).alignment = center

for i, m in enumerate(all_months_order):
    r = ctbl_row + 1 + i
    ds.cell(row=r, column=2, value=m).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    v24 = w24_map.get(m)
    v25 = w25_map.get(m)
    if v24:
        ds.cell(row=r, column=3, value=v24).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = center
    if v25:
        ds.cell(row=r, column=4, value=v25).font = data_font
    ds.cell(row=r, column=4).fill = card_fill
    ds.cell(row=r, column=4).alignment = center

c8 = LineChart()
c8.title = "Weight: 2024 vs 2025"
c8.style = 10
c8.height = 14
c8.width = 22
c8.y_axis.title = "Weight (lbs)"
compare_cats = Reference(ds, min_col=2, min_row=ctbl_row+1, max_row=ctbl_row+12)
w24_ref = Reference(ds, min_col=3, min_row=ctbl_row, max_row=ctbl_row+12)
w25_ref = Reference(ds, min_col=4, min_row=ctbl_row, max_row=ctbl_row+12)
c8.add_data(w24_ref, titles_from_data=True)
c8.add_data(w25_ref, titles_from_data=True)
c8.set_categories(compare_cats)
c8.series[0].graphicalProperties.line.width = 28000
c8.series[1].graphicalProperties.line.width = 28000
ds.add_chart(c8, f"F{ctbl_row}")

# ── Debt Reduction Chart ──
debt_row = ctbl_row + 16
ds.merge_cells(f'B{debt_row}:N{debt_row}')
ds[f'B{debt_row}'] = "DEBT REDUCTION PROGRESS (2025)"
ds[f'B{debt_row}'].font = Font(name="Arial", size=14, bold=True, color=ACCENT_GREEN)
ds[f'B{debt_row}'].fill = dark_fill

dtbl_row = debt_row + 1
ds.cell(row=dtbl_row, column=2, value="Month").font = header_font
ds.cell(row=dtbl_row, column=2).fill = green_fill
ds.cell(row=dtbl_row, column=2).alignment = center
ds.cell(row=dtbl_row, column=3, value="Debt (CAD)").font = header_font
ds.cell(row=dtbl_row, column=3).fill = green_fill
ds.cell(row=dtbl_row, column=3).alignment = center

for idx, month in enumerate(fin_2025['months']):
    r = dtbl_row + 1 + idx
    ds.cell(row=r, column=2, value=month).font = data_font
    ds.cell(row=r, column=2).fill = card_fill
    ds.cell(row=r, column=2).alignment = center
    ds.cell(row=r, column=3, value=fin_2025['debt_cad'][idx]).font = data_font
    ds.cell(row=r, column=3).fill = card_fill
    ds.cell(row=r, column=3).alignment = center
    ds.cell(row=r, column=3).number_format = '$#,##0'

debt_last = dtbl_row + len(fin_2025['months'])
c9 = BarChart()
c9.type = "col"
c9.title = "Debt Reduction (CAD)"
c9.style = 10
c9.height = 14
c9.width = 22
debt_cats = Reference(ds, min_col=2, min_row=dtbl_row+1, max_row=debt_last)
debt_vals = Reference(ds, min_col=3, min_row=dtbl_row, max_row=debt_last)
c9.add_data(debt_vals, titles_from_data=True)
c9.set_categories(debt_cats)
c9.series[0].graphicalProperties.solidFill = ACCENT_RED
ds.add_chart(c9, f"F{dtbl_row}")

# ── Set row heights for key areas ──
ds.row_dimensions[2].height = 35
ds.row_dimensions[3].height = 20
for r in [6,7,8,9]:
    ds.row_dimensions[r].height = 30

# ── Freeze panes ──
ds.freeze_panes = 'B5'

# ═══════════════════════════════════════════
# TRACKING INPUT SHEET (for continued tracking)
# ═══════════════════════════════════════════
ts = wb.create_sheet("Quick Input", 1)
ts.sheet_properties.tabColor = ACCENT_GREEN

style_range(ts, 1, 1, 60, 15, fill=dark_fill)

ts.merge_cells('B2:H2')
ts['B2'] = "MONTHLY TRACKING INPUT"
ts['B2'].font = Font(name="Arial", size=18, bold=True, color=WHITE)
ts['B2'].fill = dark_fill
ts['B2'].alignment = Alignment(horizontal="left", vertical="center")

ts.merge_cells('B3:H3')
ts['B3'] = "Enter your data here each month. Dashboard updates from original sheets."
ts['B3'].font = Font(name="Arial", size=10, color=LIGHT_GRAY)

# Column widths
ts.column_dimensions['A'].width = 3
ts.column_dimensions['B'].width = 28
ts.column_dimensions['C'].width = 15
ts.column_dimensions['D'].width = 5
ts.column_dimensions['E'].width = 28
ts.column_dimensions['F'].width = 15
ts.column_dimensions['G'].width = 5
ts.column_dimensions['H'].width = 28
ts.column_dimensions['I'].width = 15

# Health Section
ts.merge_cells('B5:C5')
ts['B5'] = "HEALTH"
ts['B5'].font = Font(name="Arial", size=12, bold=True, color=ACCENT_GREEN)
ts['B5'].fill = card_fill

health_fields = [
    "Month/Year", "Average Steps", "Total Steps", "Kms", "Current Weight",
    "Body Fat %", "Muscle Mass", "Fat free Body Weight", "Subcutaneous Fat",
    "Visceral Fat", "Bone Mass", "Protein", "Body Water %", "Metabolic Age",
    "Diet (notes)", "Exercise Routine (notes)", "Fasting done well?",
    "Alcohol / weed?", "Vitamin Status", "Milo walk routine"
]
for i, field in enumerate(health_fields):
    r = 6 + i
    cell = ts.cell(row=r, column=2, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=3)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

# Financials Section
ts.merge_cells('E5:F5')
ts['E5'] = "FINANCIALS"
ts['E5'].font = Font(name="Arial", size=12, bold=True, color=GOLD_COLOR)
ts['E5'].fill = card_fill

fin_fields = [
    "Month/Year", "Cash (CAD)", "Cash (USD)", "Debt (CAD)", "Debt (USD)",
    "Crypto (CAD)", "Crypto (USD)", "Silver (CAD)", "Silver (USD)",
    "Gold (CAD)", "Gold (USD)", "TD Stock (CAD)", "TD Stock (USD)"
]
for i, field in enumerate(fin_fields):
    r = 6 + i
    cell = ts.cell(row=r, column=5, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=6)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

# Job Section
ts.merge_cells('H5:I5')
ts['H5'] = "JOB"
ts['H5'].font = Font(name="Arial", size=12, bold=True, color=ACCENT_ORANGE)
ts['H5'].fill = card_fill

job_fields = [
    "Month/Year", "NB", "OFI", "LEI", "AHT Actual", "AHT Target",
    "Peers (notes)", "Actitud (notes)", "Career Growth (notes)"
]
for i, field in enumerate(job_fields):
    r = 6 + i
    cell = ts.cell(row=r, column=8, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=9)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

# Personal Dev Section
pd_start = 28
ts.merge_cells(f'B{pd_start}:C{pd_start}')
ts[f'B{pd_start}'] = "PERSONAL DEVELOPMENT"
ts[f'B{pd_start}'].font = Font(name="Arial", size=12, bold=True, color=ACCENT_PURPLE)
ts[f'B{pd_start}'].fill = card_fill

pd_fields = ["Month/Year", "Main Goal 1", "Main Goal 2", "Main Goal 3",
             "Secondary Goal 1", "Secondary Goal 2", "Secondary Goal 3",
             "Secondary Goal 4", "Secondary Goal 5", "Day Color (G/Y/P/B)"]
for i, field in enumerate(pd_fields):
    r = pd_start + 1 + i
    cell = ts.cell(row=r, column=2, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=3)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

# Relationship Section
ts.merge_cells(f'E{pd_start}:F{pd_start}')
ts[f'E{pd_start}'] = "RELATIONSHIP"
ts[f'E{pd_start}'].font = Font(name="Arial", size=12, bold=True, color="E91E63")
ts[f'E{pd_start}'].fill = card_fill

rel_fields = ["Month/Year", "Milestone 1", "Milestone 2", "Milestone 3",
              "Milestone 4", "Milestone 5"]
for i, field in enumerate(rel_fields):
    r = pd_start + 1 + i
    cell = ts.cell(row=r, column=5, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=6)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

# Family & Friends Section
ts.merge_cells(f'H{pd_start}:I{pd_start}')
ts[f'H{pd_start}'] = "FAMILY & FRIENDS"
ts[f'H{pd_start}'].font = Font(name="Arial", size=12, bold=True, color=ACCENT_TEAL)
ts[f'H{pd_start}'].fill = card_fill

ff_fields = ["Month/Year", "Amigos (notes)", "Familia (notes)"]
for i, field in enumerate(ff_fields):
    r = pd_start + 1 + i
    cell = ts.cell(row=r, column=8, value=field)
    cell.font = Font(name="Arial", size=10, color=WHITE)
    cell.fill = card_fill
    cell.border = thin_border
    input_cell = ts.cell(row=r, column=9)
    input_cell.fill = PatternFill("solid", fgColor="2C3E50")
    input_cell.font = Font(name="Arial", size=10, color=GOLD_COLOR)
    input_cell.border = thin_border
    input_cell.alignment = center

wb.save(out)
print(f"Dashboard saved to: {out}")
